import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = "results/reports/reporte_placas.xlsx"

HEADERS = [
    "fecha_hora",
    "placa_detectada",
    "placa_corregida",
    "tipo_detectado",
    "valido",
    "acierto",
    "metodo_color",
    "ruta_strip",
    "observacion",
]

def normalizar_placa(texto):
    texto = str(texto or "").upper()
    texto = texto.replace("-", "").replace(" ", "")
    return texto

def calcular_acierto(placa_detectada, placa_corregida):
    det = normalizar_placa(placa_detectada)
    real = normalizar_placa(placa_corregida)

    if not det or not real:
        return False

    return det == real

def crear_excel_si_no_existe():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(16, len(header) + 2)

    wb.save(EXCEL_PATH)

def guardar_reporte_excel(datos):
    crear_excel_si_no_existe()

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Resultados"]

    placa_detectada = datos.get("placa_detectada", "")
    placa_corregida = datos.get("placa_corregida", "")

    acierto = calcular_acierto(placa_detectada, placa_corregida)

    fila = {
        "fecha_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "placa_detectada": placa_detectada,
        "placa_corregida": placa_corregida,
        "tipo_detectado": datos.get("tipo_detectado", ""),
        "valido": datos.get("valido", False),
        "acierto": acierto,
        "metodo_color": datos.get("metodo_color", ""),
        "ruta_strip": datos.get("ruta_strip", ""),
        "observacion": datos.get("observacion", ""),
    }

    ws.append([fila[h] for h in HEADERS])

    # Ajuste simple de ancho
    for idx, header in enumerate(HEADERS, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 16, min(35, len(header) + 8))

    wb.save(EXCEL_PATH)

    return EXCEL_PATH
